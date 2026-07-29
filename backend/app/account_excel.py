"""会计科目 Excel:完整科目导出、二级科目导入模板、批量导入。"""
import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from . import models, subaccounts_svc

CATEGORY_CN = {"asset": "资产", "liability": "负债", "equity": "权益",
               "cost": "成本", "profit": "损益"}
DIRECTION_CN = {"debit": "借", "credit": "贷"}

THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="1F6FEB")
HEAD_FONT = Font(bold=True, color="FFFFFF")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def _style_header(ws, headers: list[str], widths: list[int]):
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, c, h)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"


def export_accounts(db: Session) -> bytes:
    """导出完整科目设置(含一二级关系)。"""
    accounts = db.scalars(
        select(models.Account).order_by(models.Account.code)
        .options(selectinload(models.Account.sub_accounts))
    ).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "会计科目表"
    headers = ["一级科目代码", "一级科目名称", "类别", "余额方向",
               "二级科目代码", "二级科目名称", "备注", "状态"]
    _style_header(ws, headers, [14, 18, 8, 10, 14, 22, 30, 8])
    r = 2
    for a in accounts:
        subs = [s for s in a.sub_accounts]
        if subs:
            for s in subs:
                _row(ws, r, [a.code, a.name, CATEGORY_CN.get(a.category, a.category),
                             DIRECTION_CN.get(a.direction, a.direction),
                             s.code, s.name, s.note,
                             "启用" if s.is_active else "停用"])
                r += 1
        else:
            _row(ws, r, [a.code, a.name, CATEGORY_CN.get(a.category, a.category),
                         DIRECTION_CN.get(a.direction, a.direction), "", "", "",
                         "启用" if a.is_active else "停用"])
            r += 1
    return _save(wb)


def _row(ws, r, values):
    for c, v in enumerate(values, 1):
        cell = ws.cell(r, c, v)
        cell.border = BORDER
        cell.alignment = LEFT if c in (2, 6, 7) else CENTER


def import_template(db: Session) -> bytes:
    """二级科目导入模板:含说明 + 一级科目参照页。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "二级科目导入"
    headers = ["一级科目代码", "一级科目名称(可留空)", "二级科目名称", "备注"]
    _style_header(ws, headers, [16, 22, 26, 30])
    # 示例行
    examples = [
        ["6602", "管理费用", "办公费", "文具、打印等"],
        ["1122", "应收账款", "XX科技有限公司", "按客户设置"],
    ]
    for i, ex in enumerate(examples, start=2):
        _row(ws, i, ex)
    ws.cell(5, 1, "说明:按行填写。一级科目代码与名称至少填一个用于匹配;二级科目代码由系统自动延续生成;已存在的同名二级会跳过。")
    ws.cell(5, 1).alignment = LEFT

    # 一级科目参照页
    ref = wb.create_sheet("一级科目参照")
    _style_header(ref, ["代码", "名称", "类别"], [12, 22, 10])
    accounts = db.scalars(select(models.Account).order_by(models.Account.code)).all()
    for i, a in enumerate(accounts, start=2):
        _row(ref, i, [a.code, a.name, CATEGORY_CN.get(a.category, a.category)])
    return _save(wb)


def import_subaccounts(db: Session, content: bytes) -> dict:
    """从 Excel 批量导入二级科目。返回统计。"""
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise ValueError("不是有效的 Excel 文件")
    ws = wb["二级科目导入"] if "二级科目导入" in wb.sheetnames else wb.worksheets[0]

    by_code = {a.code: a for a in db.scalars(select(models.Account)).all()}
    by_name = {a.name: a for a in db.scalars(select(models.Account)).all()}

    created = skipped = errors = 0
    messages: list[str] = []
    for r in range(2, ws.max_row + 1):
        code = _cell(ws, r, 1)
        pname = _cell(ws, r, 2)
        sub_name = _cell(ws, r, 3)
        note = _cell(ws, r, 4)
        if not sub_name:
            continue
        # 跳过示例/说明行
        if code and not code[0].isdigit() and not pname:
            continue
        account = by_code.get(code) or by_name.get(pname)
        if account is None:
            errors += 1
            messages.append(f"第{r}行:未找到一级科目「{code or pname}」")
            continue
        exists = db.scalar(select(models.SubAccount).where(
            models.SubAccount.account_id == account.id,
            models.SubAccount.name == sub_name))
        if exists:
            skipped += 1
            continue
        sub = subaccounts_svc.get_or_create(db, account.id, sub_name)
        if sub and note:
            sub.note = note
        created += 1
    db.commit()
    return {"created": created, "skipped": skipped, "errors": errors,
            "messages": messages[:50]}


def _cell(ws, r, c) -> str:
    v = ws.cell(r, c).value
    return str(v).strip() if v is not None else ""


def _save(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
