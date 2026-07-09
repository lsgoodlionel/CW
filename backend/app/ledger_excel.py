"""会计账簿 Excel 导出(openpyxl)。按账簿种类导出,支持导出全部六类为多 sheet。"""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from . import ledgers
from .reports_cn import Period

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
TITLE_FONT = Font(bold=True, size=13)
BOLD = Font(bold=True)
GROUP_FONT = Font(bold=True, size=11, color="1F4E79")
HEAD_FILL = PatternFill("solid", fgColor="DDEBF7")
SUM_FILL = PatternFill("solid", fgColor="FFF2CC")
MONEY = "#,##0.00"


def build_ledger_workbook(db: Session, ledger_type: str, period: Period,
                          account_code: str | None = None) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    types = list(ledgers.LEDGER_TYPES) if ledger_type == "all" else [ledger_type]
    for lt in types:
        data = ledgers.build_ledger(db, lt, period, account_code)
        _write_sheet(wb, data, period)
    if not wb.sheetnames:  # 全空兜底
        wb.create_sheet("账簿")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# 金额列索引(0-based,用于设置数字格式与右对齐),按账簿类型
def _money_cols(ledger_type: str, ncols: int) -> set[int]:
    if ledger_type in ("cash_journal", "bank_journal"):
        return {4, 5, 7}
    if ledger_type == "qty_amount":
        return {4, 6, 8}
    if ledger_type == "detail_multi":
        return set(range(3, ncols))  # 借方合计起至末尾均为金额
    return {3, 4, 6}


def _write_sheet(wb: Workbook, data: dict, period: Period):
    ledger_type = data["ledger_type"]
    columns = data["columns"]
    sub_columns = data.get("sub_columns", [])
    ncols = len(columns)
    ws = wb.create_sheet(data["title"][:31])
    last_col = get_column_letter(ncols)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"{data['title']}    {data['period_label']}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER
    money_cols = _money_cols(ledger_type, ncols)

    r = 2
    if data.get("note"):
        ws.merge_cells(f"A{r}:{last_col}{r}")
        ws.cell(r, 1, f"说明:{data['note']}").alignment = LEFT
        r += 1

    if not data["groups"]:
        ws.cell(r, 1, "本期无数据").alignment = LEFT
        _set_widths(ws, ledger_type, ncols, sub_columns)
        return

    for group in data["groups"]:
        # 科目标题行
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        gcell = ws.cell(r, 1, f"科目:{group['title']}")
        gcell.font = GROUP_FONT
        gcell.alignment = LEFT
        r += 1
        # 表头
        for c, h in enumerate(columns, 1):
            cell = ws.cell(r, c, h)
            cell.font = BOLD
            cell.alignment = CENTER
            cell.fill = HEAD_FILL
            cell.border = BORDER
        r += 1
        # 数据行
        for row in group["rows"]:
            values = ledgers.row_cells(ledger_type, row, sub_columns)
            is_sum = row.get("is_summary")
            for c, val in enumerate(values, 1):
                cell = ws.cell(r, c, val)
                cell.border = BORDER
                idx = c - 1
                if idx in money_cols:
                    cell.alignment = RIGHT
                    if val not in ("", None):
                        cell.number_format = MONEY
                elif idx == 2:
                    cell.alignment = LEFT
                else:
                    cell.alignment = CENTER
                if is_sum:
                    cell.font = BOLD
                    cell.fill = SUM_FILL
            r += 1
        r += 1  # 组间空行

    _set_widths(ws, ledger_type, ncols, sub_columns)


def _set_widths(ws, ledger_type, ncols, sub_columns):
    ws.column_dimensions["A"].width = 12   # 日期
    ws.column_dimensions["B"].width = 14   # 凭证字号
    ws.column_dimensions["C"].width = 28   # 摘要
    for c in range(4, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14
