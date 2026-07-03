"""按会小企 01/02/03 模板生成 Excel(openpyxl),供一键导出下载。

版式还原模板:标题合并、纳税人/所属期抬头、行次、列宽、边框。
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter

from sqlalchemy.orm import Session

from . import models, reports_cn
from .reports_cn import Period

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
TITLE_FONT = Font(bold=True, size=14)
BOLD = Font(bold=True)
HEAD_FILL = PatternFill("solid", fgColor="F2F2F2")


def _company(db: Session) -> models.CompanyInfo | None:
    return db.get(models.CompanyInfo, 1)


def _fmt(v) -> float | str:
    return "" if v is None else v


def build_report_workbook(db: Session, period: Period) -> bytes:
    company = _company(db)
    name = company.name if company else ""

    wb = Workbook()
    wb.remove(wb.active)
    _sheet_balance(wb, db, period, name)
    _sheet_income(wb, db, period, name)
    _sheet_cashflow(wb, db, period, name)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _header_block(ws, title: str, form_no: str, name: str, period: Period,
                  ncols: int) -> int:
    """写入标题 + 会小企表号 + 纳税人/所属期抬头,返回数据起始行(1-based)。"""
    last_col = get_column_letter(ncols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"{form_no}   单位:元"
    ws["A2"].alignment = RIGHT

    ws["A3"] = "纳税人名称"
    ws["B3"] = name
    ws["A4"] = "所属期起"
    ws["B4"] = period.cur_start.isoformat()
    half = max(3, ncols // 2)
    ws.cell(3, half).value = "所属期止"
    ws.cell(3, half + 1).value = period.cur_end.isoformat()
    ws.cell(4, half).value = "报表期间"
    ws.cell(4, half + 1).value = period.label
    for r in (3, 4):
        for c in range(1, ncols + 1):
            ws.cell(r, c).alignment = LEFT
    return 5


def _sheet_balance(wb, db, period: Period, name: str):
    ws = wb.create_sheet("资产负债表")
    data = reports_cn.balance_sheet(db, period)
    start = _header_block(
        ws, "资产负债表(适用执行小企业会计准则的企业)", "会小企01表", name, period, 8)

    headers = ["资产", "行次", "期末余额", "年初余额",
               "负债和所有者权益", "行次", "期末余额", "年初余额"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(start, c, h)
        cell.font = BOLD
        cell.alignment = CENTER
        cell.fill = HEAD_FILL
        cell.border = BORDER

    assets, rights = data["assets"], data["rights"]
    n = max(len(assets), len(rights))
    for i in range(n):
        r = start + 1 + i
        _bs_row(ws, r, 1, assets[i] if i < len(assets) else None)
        _bs_row(ws, r, 5, rights[i] if i < len(rights) else None)

    widths = [22, 6, 16, 16, 26, 6, 16, 16]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _bs_row(ws, r, col, row):
    if row is None:
        return
    label = ws.cell(r, col, row["label"])
    label.border = BORDER
    label.alignment = LEFT
    if row["style"] in ("total", "grand", "header"):
        label.font = BOLD
    lineno = ws.cell(r, col + 1, row["line"] if row["line"] else "")
    lineno.border = BORDER
    lineno.alignment = CENTER
    end = ws.cell(r, col + 2, _fmt(row.get("end")))
    beg = ws.cell(r, col + 3, _fmt(row.get("begin")))
    for cell in (end, beg):
        cell.border = BORDER
        cell.alignment = RIGHT
        cell.number_format = "#,##0.00"
        if row["style"] in ("total", "grand"):
            cell.font = BOLD


def _sheet_income(wb, db, period: Period, name: str):
    ws = wb.create_sheet("利润表")
    data = reports_cn.income_statement(db, period)
    _statement_sheet(ws, "利润表(适用执行小企业会计准则的企业)", "会小企02表",
                     name, period, data)


def _sheet_cashflow(wb, db, period: Period, name: str):
    ws = wb.create_sheet("现金流量表")
    data = reports_cn.cashflow_statement(db, period)
    _statement_sheet(ws, "现金流量表(适用执行小企业会计准则的企业)", "会小企03表",
                     name, period, data)


def _statement_sheet(ws, title, form_no, name, period: Period, data):
    """利润表 / 现金流量表通用四列版式:项目 | 行次 | 列1 | 列2。"""
    start = _header_block(ws, title, form_no, name, period, 4)
    headers = ["项目", "行次", data["col1_label"], data["col2_label"]]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(start, c, h)
        cell.font = BOLD
        cell.alignment = CENTER
        cell.fill = HEAD_FILL
        cell.border = BORDER

    for i, row in enumerate(data["rows"]):
        r = start + 1 + i
        label = ws.cell(r, 1, row["label"])
        label.border = BORDER
        label.alignment = LEFT
        if row["style"] in ("head", "total"):
            label.font = BOLD
        lineno = ws.cell(r, 2, row["line"] if row["line"] else "")
        lineno.border = BORDER
        lineno.alignment = CENTER
        c1 = ws.cell(r, 3, _fmt(row["col1"]))
        c2 = ws.cell(r, 4, _fmt(row["col2"]))
        for cell in (c1, c2):
            cell.border = BORDER
            cell.alignment = RIGHT
            cell.number_format = "#,##0.00"
            if row["style"] in ("head", "total"):
                cell.font = BOLD

    for c, w in enumerate([48, 6, 18, 18], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
